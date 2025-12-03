"""
统计服务
"""
from typing import List, Dict, Optional
from decimal import Decimal
from datetime import datetime
import csv
import io

from sqlalchemy import func
from app import db
from app.models.order import EPSOrder, EPSOrderItem, BudgetAllocation
from app.models.sku import SKU

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class StatisticsService:
    """统计服务"""
    
    @staticmethod
    def get_supplier_statistics(start_date: Optional[datetime] = None, 
                                end_date: Optional[datetime] = None) -> List[Dict]:
        """
        按供应商统计订单总额
        
        Args:
            start_date: 开始时间（可选）
            end_date: 结束时间（可选）
            
        Returns:
            供应商统计列表，每项包含 {'supplier': str, 'total_amount': Decimal, 'order_count': int}
        """
        query = db.session.query(
            EPSOrder.supplier,
            func.sum(EPSOrder.total_amount).label('total_amount'),
            func.count(EPSOrder.id).label('order_count')
        )
        
        # 应用时间范围筛选
        if start_date:
            query = query.filter(EPSOrder.created_at >= start_date)
        if end_date:
            query = query.filter(EPSOrder.created_at <= end_date)
        
        query = query.group_by(EPSOrder.supplier)
        query = query.order_by(func.sum(EPSOrder.total_amount).desc())
        
        results = query.all()
        
        return [
            {
                'supplier': row.supplier,
                'total_amount': Decimal(str(row.total_amount)) if row.total_amount else Decimal('0'),
                'order_count': row.order_count
            }
            for row in results
        ]
    
    @staticmethod
    def get_budget_statistics(budget_code: Optional[str] = None,
                              start_date: Optional[datetime] = None,
                              end_date: Optional[datetime] = None) -> List[Dict]:
        """
        按预算Code统计已使用金额和订单数量
        
        Args:
            budget_code: 预算Code（可选，如果为None则统计所有预算Code）
            start_date: 开始时间（可选）
            end_date: 结束时间（可选）
            
        Returns:
            预算统计列表，每项包含 {'budget_code': str, 'total_amount': Decimal, 'order_count': int}
        """
        query = db.session.query(
            BudgetAllocation.budget_code,
            func.sum(BudgetAllocation.amount).label('total_amount'),
            func.count(func.distinct(BudgetAllocation.order_id)).label('order_count')
        )
        
        # 如果指定了预算Code，只查询该预算Code
        if budget_code:
            query = query.filter(BudgetAllocation.budget_code == budget_code)
        
        # 应用时间范围筛选（通过关联订单表）
        if start_date or end_date:
            query = query.join(EPSOrder, BudgetAllocation.order_id == EPSOrder.id)
            if start_date:
                query = query.filter(EPSOrder.created_at >= start_date)
            if end_date:
                query = query.filter(EPSOrder.created_at <= end_date)
        
        query = query.group_by(BudgetAllocation.budget_code)
        query = query.order_by(func.sum(BudgetAllocation.amount).desc())
        
        results = query.all()
        
        return [
            {
                'budget_code': row.budget_code,
                'total_amount': Decimal(str(row.total_amount)) if row.total_amount else Decimal('0'),
                'order_count': row.order_count
            }
            for row in results
        ]
    
    @staticmethod
    def get_sku_statistics(start_date: Optional[datetime] = None,
                          end_date: Optional[datetime] = None) -> List[Dict]:
        """
        按SKU统计采购总量和总金额
        
        Args:
            start_date: 开始时间（可选）
            end_date: 结束时间（可选）
            
        Returns:
            SKU统计列表，每项包含 {'sku_id': int, 'sku_code': str, 'sku_name': str, 
                                   'total_quantity': int, 'total_amount': Decimal}
        """
        query = db.session.query(
            EPSOrderItem.sku_id,
            SKU.sku_code,
            SKU.name,
            func.sum(EPSOrderItem.quantity).label('total_quantity'),
            func.sum(EPSOrderItem.subtotal).label('total_amount')
        ).join(SKU, EPSOrderItem.sku_id == SKU.id)
        
        # 应用时间范围筛选（通过关联订单表）
        if start_date or end_date:
            query = query.join(EPSOrder, EPSOrderItem.order_id == EPSOrder.id)
            if start_date:
                query = query.filter(EPSOrder.created_at >= start_date)
            if end_date:
                query = query.filter(EPSOrder.created_at <= end_date)
        
        query = query.group_by(EPSOrderItem.sku_id, SKU.sku_code, SKU.name)
        query = query.order_by(func.sum(EPSOrderItem.subtotal).desc())
        
        results = query.all()
        
        return [
            {
                'sku_id': row.sku_id,
                'sku_code': row.sku_code,
                'sku_name': row.name,
                'total_quantity': row.total_quantity,
                'total_amount': Decimal(str(row.total_amount)) if row.total_amount else Decimal('0')
            }
            for row in results
        ]
    
    @staticmethod
    def export_to_csv(data: List[Dict], fieldnames: List[str]) -> str:
        """
        将统计数据导出为CSV格式
        
        Args:
            data: 统计数据列表
            fieldnames: 字段名列表
            
        Returns:
            CSV格式的字符串
        """
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        # 将Decimal转换为字符串以便CSV序列化
        for row in data:
            csv_row = {}
            for key, value in row.items():
                if isinstance(value, Decimal):
                    csv_row[key] = str(value)
                else:
                    csv_row[key] = value
            writer.writerow(csv_row)
        
        return output.getvalue()
    
    @staticmethod
    def import_from_csv(csv_content: str, fieldnames: List[str], 
                       decimal_fields: Optional[List[str]] = None) -> List[Dict]:
        """
        从CSV格式导入统计数据
        
        Args:
            csv_content: CSV格式的字符串
            fieldnames: 字段名列表
            decimal_fields: 需要转换为Decimal的字段列表
            
        Returns:
            统计数据列表
        """
        if decimal_fields is None:
            decimal_fields = []
        
        input_stream = io.StringIO(csv_content)
        reader = csv.DictReader(input_stream)
        
        result = []
        for row in reader:
            parsed_row = {}
            for key, value in row.items():
                if key in decimal_fields:
                    parsed_row[key] = Decimal(value)
                elif key in ['order_count', 'total_quantity', 'sku_id']:
                    # 整数字段
                    parsed_row[key] = int(value)
                else:
                    parsed_row[key] = value
            result.append(parsed_row)
        
        return result
    
    @staticmethod
    def export_to_excel(data: List[Dict], fieldnames: List[str], filename: str) -> bytes:
        """
        将统计数据导出为Excel格式
        
        Args:
            data: 统计数据列表
            fieldnames: 字段名列表
            filename: 文件名（用于工作表名称）
            
        Returns:
            Excel文件的字节内容
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Please install it to use Excel export.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistics"
        
        # 写入表头
        ws.append(fieldnames)
        
        # 写入数据
        for row in data:
            row_values = []
            for field in fieldnames:
                value = row.get(field, '')
                if isinstance(value, Decimal):
                    row_values.append(float(value))
                else:
                    row_values.append(value)
            ws.append(row_values)
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    @staticmethod
    def import_from_excel(excel_content: bytes, decimal_fields: Optional[List[str]] = None) -> List[Dict]:
        """
        从Excel格式导入统计数据
        
        Args:
            excel_content: Excel文件的字节内容
            decimal_fields: 需要转换为Decimal的字段列表
            
        Returns:
            统计数据列表
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Please install it to use Excel import.")
        
        if decimal_fields is None:
            decimal_fields = []
        
        from openpyxl import load_workbook
        
        input_stream = io.BytesIO(excel_content)
        wb = load_workbook(input_stream)
        ws = wb.active
        
        # 读取表头
        rows = list(ws.rows)
        if not rows:
            return []
        
        headers = [cell.value for cell in rows[0]]
        
        # 读取数据
        result = []
        for row in rows[1:]:
            parsed_row = {}
            for idx, cell in enumerate(row):
                if idx >= len(headers):
                    break
                
                key = headers[idx]
                value = cell.value
                
                if value is None:
                    parsed_row[key] = None
                elif key in decimal_fields:
                    parsed_row[key] = Decimal(str(value))
                elif key in ['order_count', 'total_quantity', 'sku_id']:
                    parsed_row[key] = int(value)
                else:
                    parsed_row[key] = value
            
            result.append(parsed_row)
        
        return result
